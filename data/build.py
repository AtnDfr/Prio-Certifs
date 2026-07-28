"""
Regenere index.html a partir de data/source_certifications.xlsx et data/app_template.html.

Usage:
    python3 data/build.py

Le fichier source doit contenir 2 onglets utiles :
  - "Certifications"       : catalogue (Certification, Fournisseur, Description, Difficulte)
  - "CertifesObjectifs"    : objectifs Quota/Obtenu/Restant, en 2 blocs cote a cote
                             ("Makers" et "Strategists" - le regroupement est ignore)

Le rapprochement catalogue <-> objectifs est fait par un matching automatique
(fournisseur + similarite de nom), puis corrige par 3 decisions actees avec
Antoine (Wavestone) le 2026-07-28 :
  1. "Databricks - Spark Data Engineer" (quota 3) -> ECARTE. Le nom ne
     correspond a aucune certification du catalogue de maniere fiable
     (« Data Engineer Associate » est deja utilise par une autre ligne
     d'objectif) : cette ligne d'objectif est ignoree plutot que de forcer
     un rapprochement incertain.
  2. "Azure - Designing Microsoft Azure Infrastructure Solutions (AZ-305)"
     -> "Solutions Architect Expert" (Azure). AZ-305 est le code d'examen
     officiel de cette certification Microsoft/Azure. Valide.
  3. Les 4 lignes "Microsoft 365 Copilot - ... (DW-101 a DW-104)" -> les 4
     certifications Microsoft restantes (Next Gen Productivity, Simplify
     Agent Development, Specialize Data Security Consultation, Develop
     Pro-Code Agents), par elimination et coherence thematique. Valide.
"""
import json
import re
import unicodedata
import difflib
from pathlib import Path

import openpyxl

import base64

BASE_DIR = Path(__file__).parent
SOURCE_XLSX = BASE_DIR / "source_certifications.xlsx"
TEMPLATE_HTML = BASE_DIR / "app_template.html"
LOGO_PNG = BASE_DIR / "wavestone-logo.png"
PROVIDER_LOGOS_JSON = BASE_DIR / "provider-logos.json"
LOGOS_DIR = BASE_DIR.parent / "public" / "logos"
OUTPUT_HTML = BASE_DIR.parent / "index.html"

MANUAL_OVERRIDES = {
    "Databricks - Spark Data Engineer": None,  # decision 1: ecarte
    "Azure - Designing Microsoft Azure Infrastructure Solutions (AZ-305)": ("Azure", "Solutions Architect Expert"),
    "Microsoft 365 Copilot - Copilot for Microsoft 365 Deployment Workshop (DW-101)": ("Microsoft", "Next Gen Productivity: Copilot + Agents for the Modern Enterprise"),
    "Microsoft 365 Copilot - Build and Extend AI Powered Copilot with Copilot Studio (DW-102)": ("Microsoft", "Simplify Agent Development with Copilot Studio"),
    "Microsoft 365 Copilot - Securing and Governing Copilot for Microsoft 365 with Microsoft Purview (DW-103)": ("Microsoft", "Specialize Data Security Consultation with Microsoft Purview in the Era of AI"),
    "Microsoft 365 Copilot - Extending Microsoft 365 Copilots with Pro-Code Technologies (DW-104)": ("Microsoft", "Develop Pro-Code Agents with Copilot Studio"),
}

PROVIDER_PREFIX_MAP = {
    "gcp": ["Google Cloud Platform"],
    "azure": ["Azure", "Microsoft"],
    "aure": ["Azure", "Microsoft"],
    "microsoft 365": ["Microsoft"],
    "microsoft 365 copilot": ["Microsoft"],
    "aws": ["AWS"],
    "databricks": ["Databricks"],
}


def norm(s):
    s = s.lower()
    s = re.sub(r"\(.*?\)", "", s)
    s = s.replace("-", " ")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", "-", s.lower().strip())
    return s.strip("-")


def load_catalogue(wb):
    ws = wb["Certifications"]
    rows = []
    for cert, provider, desc, difficulty in ws.iter_rows(min_row=2, values_only=True):
        if cert is None:
            continue
        rows.append({"cert": cert.strip(), "provider": provider, "desc": desc, "difficulty": difficulty})
    return rows


def load_objectives(wb):
    ws = wb["CertifésObjectifs"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        for name, quota, obtenu, restant in (row[0:4], row[5:9]):
            if name is None or name == "TOTAL":
                continue
            rows.append({"raw": name, "quota": quota, "obtenu": obtenu, "restant": restant})
    return rows


def auto_match(objective_raw, catalogue):
    prefix, _, name_part = objective_raw.partition(" - ")
    allowed = PROVIDER_PREFIX_MAP.get(prefix.strip().lower())
    name_norm = norm(name_part or objective_raw)
    best, best_score = None, -1
    for c in catalogue:
        if allowed and c["provider"] not in allowed:
            continue
        cand_norm = norm(c["cert"])
        ratio = difflib.SequenceMatcher(None, name_norm, cand_norm).ratio()
        set1, set2 = set(name_norm.split()), set(cand_norm.split())
        overlap = len(set1 & set2) / len(set1 | set2) if (set1 and set2) else 0
        score = 0.5 * ratio + 0.5 * overlap
        if score > best_score:
            best, best_score = c, score
    return best, best_score


def build_objective_lookup(catalogue, objectives):
    lookup = {}
    for obj in objectives:
        if obj["raw"] in MANUAL_OVERRIDES:
            target = MANUAL_OVERRIDES[obj["raw"]]
            if target is None:
                continue
            key = target
        else:
            match, score = auto_match(obj["raw"], catalogue)
            if not match:
                raise ValueError(f"Aucun rapprochement pour '{obj['raw']}' — ajouter une regle dans MANUAL_OVERRIDES.")
            key = (match["provider"], match["cert"].strip())
        if key in lookup:
            raise ValueError(f"Rapprochement en double sur {key} (objectif '{obj['raw']}')")
        lookup[key] = {"quota": obj["quota"], "obtenu": obj["obtenu"], "restant": obj["restant"]}
    return lookup


def build_data(catalogue, obj_lookup):
    certifications = []
    seen_ids = set()
    for c in catalogue:
        name = c["cert"].strip()
        base_id = slugify(f"{c['provider']}-{name}")
        cid, n = base_id, 2
        while cid in seen_ids:
            cid = f"{base_id}-{n}"
            n += 1
        seen_ids.add(cid)
        certifications.append({
            "id": cid,
            "name": name,
            "provider": c["provider"],
            "description": c["desc"],
            "difficulty": c["difficulty"],
            "objective": obj_lookup.get((c["provider"], name)),
        })
    providers = sorted({c["provider"] for c in certifications})
    return {"certifications": certifications, "providers": providers}


def svg_data_uri(path):
    svg_b64 = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:image/svg+xml;base64,{svg_b64}"


def build_provider_logos():
    mapping = json.loads(PROVIDER_LOGOS_JSON.read_text(encoding="utf-8"))
    logos = {}
    for provider, filename in mapping.items():
        if not filename:
            continue
        path = LOGOS_DIR / filename
        if not path.exists():
            raise ValueError(f"Logo manquant pour '{provider}': {path}")
        logos[provider] = svg_data_uri(path)
    default_uri = svg_data_uri(LOGOS_DIR / "default.svg")
    return logos, default_uri


def main():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    catalogue = load_catalogue(wb)
    objectives = load_objectives(wb)
    obj_lookup = build_objective_lookup(catalogue, objectives)
    data = build_data(catalogue, obj_lookup)

    matched = sum(1 for c in data["certifications"] if c["objective"])
    expected = len([o for o in objectives if MANUAL_OVERRIDES.get(o["raw"], "") is not None])
    print(f"certifications: {len(data['certifications'])} | avec objectif: {matched} (attendu {expected})")

    data_json = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    safe_json = data_json.replace("</", "<\\/")

    provider_logos, default_logo = build_provider_logos()
    print(f"logos fournisseurs: {len(provider_logos)}/{len(json.loads(PROVIDER_LOGOS_JSON.read_text()))} disponibles (les autres utilisent le badge par defaut)")
    logos_json = json.dumps(provider_logos, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")

    template = TEMPLATE_HTML.read_text(encoding="utf-8")
    for placeholder in ("/*__APP_DATA__*/", "/*__LOGO_DATA_URI__*/", "/*__PROVIDER_LOGOS__*/", "/*__DEFAULT_LOGO__*/"):
        if placeholder not in template:
            raise ValueError(f"Placeholder {placeholder} introuvable dans le template.")
    logo_b64 = base64.b64encode(LOGO_PNG.read_bytes()).decode("ascii")
    output = (template
              .replace("/*__APP_DATA__*/", safe_json)
              .replace("/*__LOGO_DATA_URI__*/", logo_b64)
              .replace("/*__PROVIDER_LOGOS__*/", logos_json)
              .replace("/*__DEFAULT_LOGO__*/", default_logo))
    OUTPUT_HTML.write_text(output, encoding="utf-8")
    print(f"-> {OUTPUT_HTML} ({len(output)} octets)")


if __name__ == "__main__":
    main()
